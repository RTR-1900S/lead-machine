import os
import re
import json
import time
import pandas as pd
from dotenv import load_dotenv
from google import genai
from datetime import datetime


def get_latest_leads_file(directory="output"):
    """Find the most recently created leads file (not scored)"""
    if not os.path.exists(directory):
        return None

    files = [f for f in os.listdir(directory) if f.startswith("leads_") and f.endswith(".csv") and "scored" not in f]
    if not files:
        return None

    # Sort by modification time, get the most recent
    files.sort(key=lambda f: os.path.getmtime(os.path.join(directory, f)), reverse=True)
    return os.path.join(directory, files[0])


def _pick_model(client):
    """Try models in priority order, return the first one that works."""
    model_options = [
        "gemini-2.5-flash",
        "gemini-2.0-flash",
    ]
    for model_name in model_options:
        try:
            resp = client.models.generate_content(
                model=model_name,
                contents="Say OK",
            )
            if resp.text:
                return model_name
        except Exception:
            continue
    raise Exception(f"No available Gemini models. Tried: {', '.join(model_options)}")


def _parse_json_response(text):
    """Extract JSON from a Gemini response, handling markdown fences."""
    text = text.strip()
    # Strip markdown code fences
    m = re.search(r"```(?:json)?\s*(.*?)```", text, re.DOTALL)
    if m:
        text = m.group(1).strip()
    return json.loads(text)


def score_leads(input_file=None, progress_callback=None) -> pd.DataFrame:
    """
    Loads leads from CSV, scores each with Gemini API,
    and saves results to output/leads_scored_*.csv.

    Args:
        input_file: Path to input CSV (if None, uses most recent leads file)
        progress_callback: Optional callback function(current, total) for progress tracking

    Returns:
        DataFrame with added columns: lead_score, why_good_fit, suggested_pitch
    """

    # Load environment variables
    load_dotenv("config/.env")
    api_key = os.getenv("GEMINI_API_KEY")

    if not api_key:
        raise ValueError("GEMINI_API_KEY not set in config/.env")

    # Configure Gemini (new unified SDK)
    client = genai.Client(api_key=api_key)

    # Pick a working model once (not per-lead)
    model_name = _pick_model(client)

    # Determine input file
    if input_file is None:
        input_file = get_latest_leads_file()

    if input_file is None or not os.path.exists(input_file):
        raise FileNotFoundError("No leads file found. Run scraper first.")

    df = pd.read_csv(input_file)

    # Initialize new columns
    df['lead_score'] = None
    df['why_good_fit'] = None
    df['suggested_pitch'] = None

    total_leads = len(df)

    # Score each lead
    for idx, row in df.iterrows():
        # Call progress callback if provided
        if progress_callback:
            progress_callback(idx + 1, total_leads)
        try:
            # Build a text representation of the lead
            lead_info = {
                "name": row.get("title", row.get("name", "")),
                "address": row.get("address", ""),
                "website": row.get("website", ""),
                "phone": row.get("phone", ""),
                "rating": row.get("review_rating", row.get("rating", "")),
                "review_count": row.get("review_count", ""),
                "category": row.get("category", ""),
                "emails": row.get("emails", "")
            }

            # Build the prompt for Gemini
            prompt = f"""
You are a social media marketing expert. Analyze this business lead and determine if they would be a good fit for social media management services.

Business Info:
- Name: {lead_info.get('name', 'N/A')}
- Category: {lead_info.get('category', 'N/A')}
- Address: {lead_info.get('address', 'N/A')}
- Phone: {lead_info.get('phone', 'N/A')}
- Website: {lead_info.get('website', 'N/A')}
- Rating: {lead_info.get('rating', 'N/A')}
- Review Count: {lead_info.get('review_count', 'N/A')}
- Email: {lead_info.get('emails', 'N/A')}

Score this lead (1-10) based on:
- Missing or weak Instagram/social media presence
- Low review count (indicates potential need for marketing)
- Weak or outdated website
- Likelihood they need social media management

Return ONLY valid JSON (no markdown, no explanation) in this exact format:
{{
  "lead_score": <number 1-10>,
  "why_good_fit": "<brief reason why they're a good fit>",
  "suggested_pitch": "<2-3 sentence pitch for their business>"
}}
"""

            response = client.models.generate_content(
                model=model_name,
                contents=prompt,
            )

            parsed = _parse_json_response(response.text)

            score = parsed.get('lead_score', 0)
            # Clamp score to 1-10
            score = max(1, min(10, int(score))) if score else 0
            df.at[idx, 'lead_score'] = score
            df.at[idx, 'why_good_fit'] = parsed.get('why_good_fit', '')
            df.at[idx, 'suggested_pitch'] = parsed.get('suggested_pitch', '')

        except json.JSONDecodeError:
            print(f"⚠️  JSON parse error for row {idx}")
            df.at[idx, 'lead_score'] = 0
            df.at[idx, 'why_good_fit'] = "Error parsing response"
            df.at[idx, 'suggested_pitch'] = "Could not generate pitch"

        except Exception as e:
            print(f"❌ Error scoring row {idx}: {e}")
            df.at[idx, 'lead_score'] = 0
            df.at[idx, 'why_good_fit'] = f"API Error: {str(e)[:50]}"
            df.at[idx, 'suggested_pitch'] = "Retry or check API key"

        # Small delay to avoid rate limits
        time.sleep(0.5)

    # Save scored leads with timestamp
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    csv_output_file = os.path.join("output", f"leads_scored_{timestamp}.csv")
    df.to_csv(csv_output_file, index=False)

    # Also save Excel version with formatting
    try:
        excel_output_file = os.path.join("output", f"leads_scored_{timestamp}.xlsx")
        save_scored_leads_excel(df, excel_output_file)
    except Exception as e:
        print(f"Warning: Could not save Excel file: {e}")

    return df


def save_scored_leads_excel(df: pd.DataFrame, output_path: str):
    """
    Save scored leads to Excel with professional formatting.

    Args:
        df: DataFrame with scored leads
        output_path: Path to save Excel file
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Scored Leads"

        # Add header row
        headers = ["Name", "Category", "Phone", "Email", "Website", "Rating", "Review Count", "Lead Score", "Why Good Fit", "Suggested Pitch"]
        ws.append(headers)

        # Format header row
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        # Add data rows
        for idx, row in df.iterrows():
            values = [
                row.get("title", row.get("name", "")),
                row.get("category", ""),
                row.get("phone", ""),
                row.get("emails", ""),
                row.get("website", ""),
                row.get("review_rating", row.get("rating", "")),
                row.get("review_count", ""),
                row.get("lead_score", ""),
                row.get("why_good_fit", ""),
                row.get("suggested_pitch", "")
            ]
            ws.append(values)

            # Format data rows
            for cell_idx, cell in enumerate(ws[idx + 2], 1):
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

                # Color code the score column
                if cell_idx == 8:  # lead_score column
                    try:
                        score = float(cell.value) if cell.value else 0
                        if score >= 7:
                            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")  # Green
                            cell.font = Font(color="FFFFFF", bold=True)
                        elif score >= 5:
                            cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # Orange
                            cell.font = Font(color="000000", bold=True)
                        else:
                            cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Red
                            cell.font = Font(color="FFFFFF", bold=True)
                    except (ValueError, TypeError):
                        pass

        # Adjust column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 25
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["H"].width = 12
        ws.column_dimensions["I"].width = 25
        ws.column_dimensions["J"].width = 30

        # Set row height for header
        ws.row_dimensions[1].height = 30

        # Save workbook
        wb.save(output_path)

    except ImportError:
        # openpyxl not available, just skip Excel export
        pass
